#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Authors: Bryan Kyung and Katie Kim
"""
Score the human-coded discriminatory gold set.

Reads the completed coding workbook (Discriminatory_Coding_Sheet.xlsx, 'Coding' sheet)
plus the hidden machine key (coding_key.csv), then:

  1. Reports inter-coder agreement (raw % and Cohen's kappa) on the two coder columns.
  2. Builds the human gold label by reconciliation:
       - if the two coders agree, that value is the gold label;
       - disagreements are listed for the coders to resolve (edit a 'final' value in
         reconcile.csv, or pass --tiebreak lenient/strict to auto-resolve: lenient=1 if
         either coded 1, strict=1 only if both coded 1).
  3. Computes discriminatory precision / recall / F1 for each classifier
     (context-aware lexicon, zero-shot DeBERTa) against the human gold.

Because the sample is ENRICHED (machine-flagged sentences are over-sampled relative to
the corpus), every corpus-level count is inverse-probability weighted by samp_weight so
the precision/recall/F1 estimate the values that would obtain on the full 24,546-sentence
corpus, not on the (positive-heavy) sample.

USAGE
    pip install openpyxl
    python score_discriminatory_gold.py \
        --coded Discriminatory_Coding_Sheet.xlsx \
        --key   coding_key.csv \
        --tiebreak report          # or: lenient | strict
"""
import argparse, csv, math
from collections import defaultdict

def read_coding(path):
    from openpyxl import load_workbook
    wb=load_workbook(path,data_only=True); ws=wb["Coding"]
    rows=list(ws.iter_rows(values_only=True)); head=[str(h).strip() if h else '' for h in rows[0]]
    idx={h:i for i,h in enumerate(head)}
    out=[]
    for r in rows[1:]:
        if r[idx['code_id']] is None: continue
        def g(col):
            v=r[idx[col]]
            return None if v is None or str(v).strip()=='' else int(float(v))
        out.append(dict(code_id=str(r[idx['code_id']]).strip(),
                        c1=g('coder1_disc'), c2=g('coder2_disc')))
    return out

def cohen_kappa(pairs):
    # pairs: list of (a,b) in {0,1}
    n=len(pairs)
    if n==0: return float('nan')
    po=sum(1 for a,b in pairs if a==b)/n
    pa1=sum(1 for a,b in pairs if a==1)/n; pb1=sum(1 for a,b in pairs if b==1)/n
    pe=pa1*pb1+(1-pa1)*(1-pb1)
    return (po-pe)/(1-pe) if pe!=1 else float('nan')

def main():
    ap=argparse.ArgumentParser()
    ap.add_argument("--coded",default="Discriminatory_Coding_Sheet.xlsx")
    ap.add_argument("--key",default="coding_key.csv")
    ap.add_argument("--tiebreak",default="report",choices=["report","lenient","strict"])
    a=ap.parse_args()

    key={r['code_id']:r for r in csv.DictReader(open(a.key,encoding='utf-8'))}
    coded=read_coding(a.coded)

    both=[(r['c1'],r['c2']) for r in coded if r['c1'] is not None and r['c2'] is not None]
    n_both=len(both)
    print(f"\nRows coded by both coders: {n_both} / {len(coded)}")
    if n_both:
        agree=sum(1 for a_,b_ in both if a_==b_)
        print(f"Raw agreement : {agree}/{n_both} = {agree/n_both:.3f}")
        print(f"Cohen's kappa : {cohen_kappa(both):.3f}")

    # build gold + disagreement list
    gold={}; disagree=[]
    for r in coded:
        c1,c2=r['c1'],r['c2']
        if c1 is None and c2 is None: continue
        if c1 is None or c2 is None:
            gold[r['code_id']]=c1 if c1 is not None else c2
        elif c1==c2:
            gold[r['code_id']]=c1
        else:
            disagree.append(r['code_id'])
            if a.tiebreak=="lenient": gold[r['code_id']]=1
            elif a.tiebreak=="strict": gold[r['code_id']]=0
    if disagree and a.tiebreak=="report":
        with open("reconcile.csv","w",newline='',encoding='utf-8') as f:
            w=csv.writer(f); w.writerow(['code_id','note_resolve_here'])
            for cid in disagree: w.writerow([cid,''])
        print(f"\n{len(disagree)} disagreements written to reconcile.csv "
              f"(resolve them, or rerun with --tiebreak lenient/strict). "
              f"Scores below EXCLUDE unresolved disagreements.")

    # weighted P/R/F1 for each classifier
    print("\n=== DISCRIMINATORY VALIDATION vs. human gold (inverse-probability weighted) ===")
    print(f"{'classifier':18s} {'TP':>7} {'FP':>7} {'FN':>7} {'Prec':>6} {'Recall':>7} {'F1':>6}")
    for clf,col in [("context-lexicon","lex_disc"),("zero-shot DeBERTa","zs_disc")]:
        tp=fp=fn=0.0
        for cid,g in gold.items():
            if cid not in key: continue
            w=float(key[cid]['samp_weight']); pred=int(key[cid][col])
            if pred==1 and g==1: tp+=w
            elif pred==1 and g==0: fp+=w
            elif pred==0 and g==1: fn+=w
        P=tp/(tp+fp) if tp+fp else float('nan')
        R=tp/(tp+fn) if tp+fn else float('nan')
        F=2*P*R/(P+R) if (P==P and R==R and P+R>0) else float('nan')
        print(f"{clf:18s} {tp:7.0f} {fp:7.0f} {fn:7.0f} {P:6.2f} {R:7.2f} {F:6.2f}")

    # raw (unweighted) human prevalence in the coded sample
    npos=sum(1 for g in gold.values() if g==1)
    print(f"\nHuman-labeled discriminatory in coded sample: {npos}/{len(gold)} "
          f"({(npos/len(gold)*100 if gold else 0):.1f}%).")
    # estimated corpus prevalence via weights
    wpos=sum(float(key[c]['samp_weight']) for c,g in gold.items() if g==1 and c in key)
    wall=sum(float(key[c]['samp_weight']) for c in gold if c in key)
    if wall: print(f"Inverse-probability-weighted corpus prevalence estimate: {wpos/wall*100:.2f}% "
                   f"(~{wpos:.0f} of ~{wall:.0f} sentences).")

if __name__=="__main__":
    main()
