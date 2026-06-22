#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Authors: Bryan Kyung and Katie Kim
"""
Train and cross-validate a supervised classifier for the AGGRESSIVE and INFLAMMATORY
categories, using the human labels in Aggressive_Inflammatory_Coding_Sheet.xlsx.

This mirrors the discriminatory supervised pipeline (TF-IDF + logistic regression,
5-fold stratified cross-validation, averaged over 10 random splits). It fills in the
two '—' cells in the category-level validation table (Table 11) once the sheet is coded.

Run:  pip install openpyxl numpy
      python train_supervised_aggr_infl.py --coded Aggressive_Inflammatory_Coding_Sheet.xlsx
"""
import argparse, re
import numpy as np
from collections import Counter

def load(path):
    from openpyxl import load_workbook
    ws=load_workbook(path,data_only=True)["Coding"]
    rows=list(ws.iter_rows(values_only=True)); head=[str(h).strip() if h else '' for h in rows[0]]
    ix={h:i for i,h in enumerate(head)}
    S=[]; A1=[];A2=[];I1=[];I2=[]
    def g(r,c):
        v=r[ix[c]]; return None if v is None or str(v).strip()=='' else int(float(v))
    for r in rows[1:]:
        if r[ix['code_id']] is None: continue
        S.append(str(r[ix['sentence']]))
        A1.append(g(r,'aggr_coder1'));A2.append(g(r,'aggr_coder2'))
        I1.append(g(r,'infl_coder1'));I2.append(g(r,'infl_coder2'))
    return S,(A1,A2),(I1,I2)

def kappa(a,b):
    n=len(a); po=sum(x==y for x,y in zip(a,b))/n
    pa=sum(a)/n; pb=sum(b)/n; pe=pa*pb+(1-pa)*(1-pb)
    return (po-pe)/(1-pe) if pe!=1 else float('nan')

def tok(s):
    t=re.findall(r"[a-z']+",str(s).lower()); return t+[t[i]+'_'+t[i+1] for i in range(len(t)-1)]
def vocab(docs,minc=2):
    c=Counter()
    for d in docs: c.update(set(tok(d)))
    return {w:i for i,w in enumerate([w for w,n in c.items() if n>=minc])}
def idf(docs,v):
    df=np.zeros(len(v))
    for d in docs:
        for w in set(tok(d)):
            if w in v: df[v[w]]+=1
    return np.log((1+len(docs))/(1+df))+1
def feats(docs,v, id_):
    X=np.zeros((len(docs),len(v)))
    for r,d in enumerate(docs):
        for w in tok(d):
            if w in v: X[r,v[w]]+=1
    X=np.log1p(X)*id_; n=np.linalg.norm(X,axis=1,keepdims=True); n[n==0]=1; return X/n
def lr(X,y,l2=1.0,it=600,lr_=0.5):
    n,p=X.shape; w=np.zeros(p); b=0.0
    cw=np.where(y==1,n/(2*max(y.sum(),1)),n/(2*max((1-y).sum(),1)))
    for _ in range(it):
        pr=1/(1+np.exp(-(X@w+b))); g=(pr-y)*cw; w-=lr_*(X.T@g/n+l2*w/n); b-=lr_*g.mean()
    return w,b

def cv_f1(S,y,seeds=10,k=5):
    def folds(y,rng):
        i0,i1=np.where(y==0)[0],np.where(y==1)[0]; rng.shuffle(i0); rng.shuffle(i1)
        F=[[] for _ in range(k)]
        for i,ix in enumerate(i1):F[i%k].append(ix)
        for i,ix in enumerate(i0):F[i%k].append(ix)
        return [np.array(f) for f in F]
    out=[]
    for s in range(seeds):
        rng=np.random.RandomState(s); fs=folds(y,rng); oof=np.zeros(len(y))
        for f in range(k):
            te=fs[f]; tr=np.concatenate([fs[j] for j in range(k) if j!=f])
            v=vocab([S[i] for i in tr]); id_=idf([S[i] for i in tr],v)
            w,b=lr(feats([S[i] for i in tr],v,id_),y[tr])
            oof[te]=1/(1+np.exp(-(feats([S[i] for i in te],v,id_)@w+b)))
        p=(oof>=0.5).astype(int)
        tp=int(((p==1)&(y==1)).sum());fp=int(((p==1)&(y==0)).sum());fn=int(((p==0)&(y==1)).sum())
        P=tp/(tp+fp) if tp+fp else 0;R=tp/(tp+fn) if tp+fn else 0
        out.append(2*P*R/(P+R) if P+R else 0)
    return float(np.mean(out)),float(np.std(out))

def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--coded",default="Aggressive_Inflammatory_Coding_Sheet.xlsx")
    a=ap.parse_args()
    S,(A1,A2),(I1,I2)=load(a.coded)
    for name,(c1,c2) in [("AGGRESSIVE",(A1,A2)),("INFLAMMATORY",(I1,I2))]:
        if any(x is None for x in c1+c2):
            print(f"{name}: not fully coded yet — fill both coder columns, then rerun."); continue
        c1=np.array(c1);c2=np.array(c2)
        print(f"\n=== {name} ===")
        print(f"  inter-coder: raw agreement {np.mean(c1==c2):.3f}, Cohen's kappa {kappa(c1,c2):.3f}")
        for rule,y in [("lenient",(c1|c2).astype(int)),("strict",(c1&c2).astype(int))]:
            f,sd=cv_f1(S,y)
            print(f"  [{rule}] positives={int(y.sum())}/{len(y)}  cross-validated F1 = {f:.2f} ± {sd:.2f}")

if __name__=="__main__":
    main()
